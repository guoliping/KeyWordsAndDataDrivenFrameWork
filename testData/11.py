from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Border,Side,Font

workbook = openpyxl.load_workbook(u"h:\\DataDrivenFrameWork\\testData\\126邮箱联系人.xlsx") #导入工作表

a_sheet=workbook.get_sheet_by_name(u'联系人')
print(a_sheet.title)
sheet=workbook.active
b4=sheet['B4']
print(f'({b4.column},{b4.row}) is {b4.value}')
b4_too=sheet.cell(row=4,column=2)
print(b4_too.value)
#sheetnames =workbook.get_sheet_name() #获得表单名字
#sheet = workbook.get_sheet_by_name(sheetnames[0]) #从工作表中提取某一表单
print(workbook.get_active_sheet().title)
for i in workbook.get_sheet_names():
    print (i.title)

#for row_cell in sheet['A1':'B3']

for row in sheet.rows:
    for cell in row:
        print(cell.value)
for column in sheet.columns:
    for cell in column:
        print(cell.value)